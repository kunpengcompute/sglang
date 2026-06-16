# Copyright 2026 Huawei Technologies Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

"""
Collect execution time of specific functions from Torch Profiler JSON,
and output statistics to an Excel file (multi-sheet).
Features:
- Robust parsing: skip malformed events
- Convert dur from microseconds to milliseconds
- Substring matching on raw_name
- Generate Excel file (profiler_stats.xlsx), each run adds a new sheet named by timestamp
- CSV output mode as fallback only
"""

import json
import sys
import argparse
import csv
import os
import re
from datetime import datetime

# ===== Target function names for JSON search =====
TARGET_NAMES = [
    "recv_requests",
    "get_next_batch_to_run",
    "init_forward_metadata",
    "VocabParallelEmbedding_0",
    "_scattered_to_tp_attn_full",
    "forward_prepare",
    "forward_core",
    "_all_reduce_and_layernorm",
    "deepseek_v2.py(255): forward",
    "_scatter_hidden_states_and_residual",
    "deepseek_v2.py(323): forward",
    "topk.py(1246): select_experts",
    "kunpeng.py(479): dispatch",
    "layer.py(680): run_moe_core",
    "kunpeng.py(580): combine",
    "logits_processor.py(285): forward",
    "model_runner.py(3341): sample",
    "process_batch_result",
]

# ===== Display names for Excel output (1:1 mapping with TARGET_NAMES) =====
DISPLAY_NAMES = [
    "recv_requests",
    "get_next_batch_to_run",
    "init_forward_metadata",
    "vocab_parallel_embedding",
    "prepare_attn(allgather)",
    "forward_prepare",
    "forward_core",
    "prepare_mlp(dense)",
    "forward_mlp",
    "prepare_mlp(sparse)",
    "moe_gate",
    "topk",
    "moe_dispatch",
    "run_moe_core",
    "moe_combine",
    "logits_processor",
    "sampler",
    "process_batch_result",
]
# ===================================================================

FRONT_COUNT = 4  # First FRONT_COUNT items use total=avg_ms


def iter_events_robust(file_path):
    """Safely iterate JSON events, skipping corrupted parts"""
    with open(file_path, "r", encoding="utf-8") as f:
        content = f.read()
    decoder = json.JSONDecoder()
    match = re.search(r'"traceEvents"\s*:\s*\[', content)
    if match:
        start = match.end()
    else:
        start = content.find("[")
        if start == -1:
            raise ValueError("Cannot find traceEvents array or root array")
        start += 1
    idx = start
    while idx < len(content):
        idx = content.find("{", idx)
        if idx == -1:
            break
        try:
            obj, end = decoder.raw_decode(content, idx)
            yield obj
            idx = end
        except json.JSONDecodeError as e:
            idx = e.pos + 1


def parse_args():
    parser = argparse.ArgumentParser(
        description="Collect dur statistics of specific functions from Torch Profiler"
    )
    parser.add_argument("file", help="Path to the JSON file")
    return parser.parse_args()


def main():
    args = parse_args()
    file_path = args.file

    # Collect statistics
    stats = {name: [] for name in TARGET_NAMES}
    try:
        for ev in iter_events_robust(file_path):
            if ev.get("ph") != "X":
                continue
            raw_name = ev.get("name", "")
            dur_us = ev.get("dur")
            if dur_us is None:
                continue
            for target in TARGET_NAMES:
                if target in raw_name:
                    dur_ms = float(dur_us) / 1000.0
                    stats[target].append(dur_ms)
                    break
    except Exception as e:
        print(f"Parse error: {e}", file=sys.stderr)
        sys.exit(1)

    # Base count for normalization
    base_durations = stats.get("VocabParallelEmbedding_0", [])
    base_count = len(base_durations)
    if base_count == 0:
        print(
            "Warning: 'VocabParallelEmbedding_0' event not found, normalized total will be 0",
            file=sys.stderr,
        )

    # Build result list
    results = []
    print(f"File: {file_path}")
    print("-" * 60)
    for i, name in enumerate(TARGET_NAMES):
        display_name = DISPLAY_NAMES[i]
        durations = stats[name]
        count = len(durations)
        if count == 0:
            print(f"Function '{name}': no events found")
            total = None if i < FRONT_COUNT else 0.0
            results.append(
                {
                    "function": display_name,
                    "count": 0,
                    "min_ms": None,
                    "max_ms": None,
                    "avg_ms": None,
                    "total": total,
                }
            )
            continue

        min_val = min(durations)
        max_val = max(durations)
        avg_val = sum(durations) / count

        if i < FRONT_COUNT:
            total = avg_val
        else:
            total = (avg_val * count) / base_count if base_count > 0 else 0.0

        print(f"Function '{name}':")
        print(f"  Count:  {count}")
        print(f"  Min:    {min_val:.3f} ms")
        print(f"  Max:    {max_val:.3f} ms")
        print(f"  Avg:    {avg_val:.3f} ms")
        print(f"  total:  {total:.3f} ms" if total is not None else "  total:  N/A")
        print()

        results.append(
            {
                "function": display_name,
                "count": count,
                "min_ms": min_val,
                "max_ms": max_val,
                "avg_ms": avg_val,
                "total": total,
            }
        )

    total_sum = sum(r["total"] for r in results if r["total"] is not None)

    # Prepare Excel output
    excel_file = "prof_stats.xlsx"
    sheet_name = datetime.now().strftime("%Y-%m-%d %H-%M-%S")

    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: openpyxl is required for Excel multi-sheet output.")
        print("Please run: pip install openpyxl")
        sys.exit(1)

    # Load or create workbook
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Create new sheet (insert at the left so newest is first)
    ws = wb.create_sheet(title=sheet_name, index=0)

    # Write header row
    headers = ["function", "count", "min_ms", "max_ms", "avg_ms", "total"]
    ws.append(headers)

    # Write data rows
    for r in results:
        row = [
            r["function"],
            r["count"],
            round(r["min_ms"], 3) if r["min_ms"] is not None else "",
            round(r["max_ms"], 3) if r["max_ms"] is not None else "",
            round(r["avg_ms"], 3) if r["avg_ms"] is not None else "",
            round(r["total"], 3) if r["total"] is not None else "",
        ]
        ws.append(row)

    # Append summary row
    ws.append(["Total", "", "", "", "", round(total_sum, 3)])

    # Auto-adjust column width (optional)
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 18

    # Save workbook
    wb.save(excel_file)
    print(f"Results appended to {excel_file}, sheet: {sheet_name}")


if __name__ == "__main__":
    main()
